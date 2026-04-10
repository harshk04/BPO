import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from pathlib import Path
from threading import Lock, local
from types import SimpleNamespace
from typing import Dict, List, Optional, Set, Tuple

from dotenv import load_dotenv
from google import genai
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from excel_loan_calculator import LoanOutputs, calculate
from maingem import ApiRateLimiter, compute_insurance_rate, get_image_paths, process_single_image
from roboflow_inference import (
    DEFAULT_CONFIDENCE,
    DEFAULT_OVERLAP,
    ensure_dependencies as ensure_roboflow_dependencies,
    load_model as load_roboflow_model,
    load_settings as load_roboflow_settings,
    run_inference,
)
from run_logger import RunLogger

INPUT_IMAGES_FOLDER = "images"
ROBOFLOW_OUTPUT_ROOT = "outputs"
ROBOFLOW_CONFIDENCE = DEFAULT_CONFIDENCE
ROBOFLOW_OVERLAP = DEFAULT_OVERLAP
OUTPUT_EXCEL_PATH = "final_loan_outputs.xlsx"
LOGS_DIR = "logs"
SHEET_NAME = "Final_Output"
DEFAULT_CROP_WORKERS = 4
DEFAULT_GEMINI_MAX_REQUESTS_PER_MINUTE = 0

EXCEL_HEADERS = [
    "sample no . , record no,",
    "Customer Reference Number",
    "Customer name",
    "City,State",
    "Purchase Value AND Down payment",
    "Loan Period AND Annual Interest",
    "Guarantor name",
    "Guarantor reference number",
    "Loan Amount AND Principal",
    "Total Interest for Loan Period AND Property Insurance Per Month",
]

COMMA_PADDED_SEPARATOR = "  ,  "
_WORKER_STATE = local()


class GeminiKeyManager:
    def __init__(self, api_keys: List[str], logger: RunLogger) -> None:
        if not api_keys:
            raise ValueError("At least one Gemini API key is required.")
        self._keys = api_keys
        self._logger = logger
        self._active_index = 0
        self._lock = Lock()

    @property
    def key_count(self) -> int:
        return len(self._keys)

    def get_active(self) -> Tuple[int, str]:
        with self._lock:
            return self._active_index, self._keys[self._active_index]

    def rotate_after_failure(
        self,
        failed_index: int,
        crop_name: str,
        error: Exception,
    ) -> Tuple[int, str, bool]:
        reason = compact_error_message(error)
        with self._lock:
            active_before = self._active_index
            if active_before != failed_index:
                current_index = self._active_index
                current_key = self._keys[current_index]
                self._logger.log(
                    f"gemini_key_rotation_skip crop={crop_name} "
                    f"failed_index={failed_index + 1} current_index={current_index + 1}"
                )
                return current_index, current_key, False

            next_index = (failed_index + 1) % len(self._keys)
            if next_index == failed_index:
                self._logger.log(
                    f"gemini_key_rotation_unavailable crop={crop_name} "
                    f"key_index={failed_index + 1} reason={reason}"
                )
                return failed_index, self._keys[failed_index], False

            self._active_index = next_index
            failed_masked = mask_api_key(self._keys[failed_index])
            next_masked = mask_api_key(self._keys[next_index])
            self._logger.log(
                f"gemini_key_rotated crop={crop_name} "
                f"from_key_index={failed_index + 1} from_key={failed_masked} "
                f"to_key_index={next_index + 1} to_key={next_masked} reason={reason}"
            )
            return next_index, self._keys[next_index], True


def extract_decimal(value: str, field_name: str) -> Decimal:
    text = str(value or "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        raise ValueError(f"Could not parse numeric value for {field_name}: {value!r}")
    try:
        return Decimal(match.group(0))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value for {field_name}: {value!r}") from exc


def decimal_to_plain(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def apply_padded_comma_spacing(text: str) -> str:
    return text.replace(",", COMMA_PADDED_SEPARATOR)


def format_currency(value: Decimal, spaces_after_dollar: int = 2) -> str:
    grouped = apply_padded_comma_spacing(f"{value:,.2f}")
    return f"$" + (" " * spaces_after_dollar) + grouped


def format_currency_trimmed(value: Decimal, spaces_after_dollar: int = 2) -> str:
    text = f"{value:,.2f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    grouped = apply_padded_comma_spacing(text)
    return f"$" + (" " * spaces_after_dollar) + grouped


def format_currency_first_digit_group(value: Decimal, spaces_after_dollar: int = 2) -> str:
    text = f"{value:.2f}"
    integer_part, fractional_part = text.split(".")
    if len(integer_part) > 1:
        grouped_integer = f"{integer_part[:1]},{integer_part[1:]}"
    else:
        grouped_integer = integer_part
    grouped = apply_padded_comma_spacing(f"{grouped_integer}.{fractional_part}")
    return f"$" + (" " * spaces_after_dollar) + grouped


def format_percent(raw_value: str) -> str:
    value = extract_decimal(raw_value, "percent_value")
    return f"{decimal_to_plain(value)} %"


def parse_percent_for_calculator(raw_value: str, field_name: str) -> Decimal:
    value = extract_decimal(raw_value, field_name)
    if "%" in str(raw_value):
        return value / Decimal("100")
    return value


def format_upper_words(raw_value: str, spaces_between_words: int) -> str:
    tokens = [token for token in str(raw_value or "").split() if token]
    separator = " " * spaces_between_words
    return separator.join(token.upper() for token in tokens)


def format_city_state(city: str, state: str) -> str:
    city_clean = " ".join(str(city or "").split()).upper()
    state_clean = " ".join(str(state or "").split()).upper()
    return f"{city_clean} , {state_clean}"


def format_loan_period_and_interest(record: dict) -> str:
    years = extract_decimal(record["loan_period_in_year"], "loan_period_in_year")
    years_text = f"{int(years)} YEARS"
    annual_interest_text = format_percent(record["annual_interest"])
    return f"{years_text} AND {annual_interest_text}"


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 18), 80)


def ensure_workbook(path: str):
    excel_path = Path(path)
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        existing_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(EXCEL_HEADERS))]
        if existing_headers != EXCEL_HEADERS:
            ws.delete_rows(1, ws.max_row)
            ws.append(EXCEL_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(EXCEL_HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return wb, ws


def append_row_to_excel(path: str, row: List[str]) -> None:
    wb, ws = ensure_workbook(path)
    ws.append(row)

    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_columns(ws)
    wb.save(path)


def calculate_outputs(record: dict) -> Tuple[LoanOutputs, bool]:
    insurance_rate = compute_insurance_rate(record["discount_percent"], record["loan_period_in_year"])
    insurance_na = insurance_rate == "NA"

    insurance_rate_percent = (
        Decimal("0")
        if insurance_na
        else parse_percent_for_calculator(insurance_rate, "insurance_rate")
    )

    # To match your requested output format, map as follows:
    # reduction_percent <- reduction_percent_value_down_payment
    # down_payment_percent <- discount_percent
    result = calculate(
        purchase_value=extract_decimal(record["purchase_value"], "purchase_value"),
        reduction_percent=parse_percent_for_calculator(
            record["reduction_percent_value_down_payment"],
            "reduction_percent_value_down_payment",
        ),
        down_payment_percent=parse_percent_for_calculator(
            record["discount_percent"],
            "discount_percent",
        ),
        loan_period_years=extract_decimal(record["loan_period_in_year"], "loan_period_in_year"),
        monthly_principal_reduction_percent=parse_percent_for_calculator(
            record["monthly_principal_reduction_percent"],
            "monthly_principal_reduction_percent",
        ),
        interest_rate_percent=parse_percent_for_calculator(
            record["int_rate_percent"],
            "int_rate_percent",
        ),
        total_interest_reduction_percent=parse_percent_for_calculator(
            record["total_interest_reduction"],
            "total_interest_reduction",
        ),
        insurance_rate_percent=insurance_rate_percent,
    )

    return result, insurance_na


def build_excel_row(record: dict, result: LoanOutputs, insurance_na: bool) -> List[str]:
    total_interest_and_insurance = (
        f"{format_currency_trimmed(result.total_interest_after_reduction, spaces_after_dollar=3)} AND NA"
        if insurance_na
        else (
            f"{format_currency_trimmed(result.total_interest_after_reduction, spaces_after_dollar=3)} AND "
            f"{format_currency_first_digit_group(result.insurance_amount_per_month, spaces_after_dollar=2)}"
        )
    )

    return [
        f"SAMPLE NO. {record['sample_no']} ,RECORD.{record['record_no']}",
        format_upper_words(record["customer_reference_number"], spaces_between_words=3),
        format_upper_words(record["customer_name"], spaces_between_words=2),
        format_city_state(record["city"], record["state"]),
        (
            f"{format_currency_trimmed(result.purchase_value_reduction_amount, spaces_after_dollar=2)} AND "
            f"{format_percent(record['discount_percent'])}"
        ),
        format_loan_period_and_interest(record),
        format_upper_words(record["guarantor_name"], spaces_between_words=2),
        format_upper_words(record["guarantor_reference_number"], spaces_between_words=3),
        (
            f"{format_currency_trimmed(result.loan_amount, spaces_after_dollar=2)} AND "
            f"{format_currency_trimmed(result.principal_after_monthly_reduction, spaces_after_dollar=2)}"
        ),
        total_interest_and_insurance,
    ]


def build_roboflow_settings():
    args = SimpleNamespace(confidence=ROBOFLOW_CONFIDENCE, overlap=ROBOFLOW_OVERLAP)
    return load_roboflow_settings(args)


def parse_int_env(name: str, default: int, minimum: int = 0) -> int:
    raw_value = os.getenv(name, "").strip()
    if not raw_value:
        return default
    try:
        value = int(raw_value)
    except ValueError as exc:
        raise RuntimeError(f"{name} must be an integer. Got: {raw_value!r}") from exc
    if value < minimum:
        raise RuntimeError(f"{name} must be >= {minimum}. Got: {value}")
    return value


def compact_error_message(error: Exception) -> str:
    text = " ".join(str(error).split())
    if len(text) > 300:
        return text[:297] + "..."
    return text


def mask_api_key(api_key: str) -> str:
    key = (api_key or "").strip()
    if not key:
        return "<empty>"
    if len(key) <= 8:
        return f"***{key[-2:]}"
    return f"{key[:4]}...{key[-4:]}"


def parse_gemini_api_keys() -> List[str]:
    discovered: List[Tuple[int, str]] = []
    indexed_pattern = re.compile(r"^GEMINI_API_KEY(\d+)$")

    for env_key, raw_value in os.environ.items():
        match = indexed_pattern.match(env_key)
        if not match:
            continue
        key_value = raw_value.strip()
        if not key_value:
            continue
        discovered.append((int(match.group(1)), key_value))

    discovered.sort(key=lambda item: item[0])
    ordered_keys = [value for _, value in discovered]

    fallback_key = os.getenv("GEMINI_API_KEY", "").strip()
    if fallback_key:
        ordered_keys.insert(0, fallback_key)

    deduped: List[str] = []
    seen: Set[str] = set()
    for key in ordered_keys:
        if key in seen:
            continue
        seen.add(key)
        deduped.append(key)
    return deduped


def get_worker_client(api_key: str) -> genai.Client:
    clients = getattr(_WORKER_STATE, "gemini_clients", None)
    if clients is None:
        clients = {}
        _WORKER_STATE.gemini_clients = clients
    client = clients.get(api_key)
    if client is None:
        client = genai.Client(api_key=api_key)
        clients[api_key] = client
    return client


def process_crop_record(
    crop_path: Path,
    key_manager: GeminiKeyManager,
    model: str,
    logger: RunLogger,
    rate_limiter: ApiRateLimiter,
) -> Dict[str, str]:
    attempted_indices: Set[int] = set()
    last_error: Optional[Exception] = None

    while len(attempted_indices) < key_manager.key_count:
        key_index, api_key = key_manager.get_active()
        attempted_indices.add(key_index)
        logger.log(
            f"gemini_key_attempt crop={crop_path.name} "
            f"key_index={key_index + 1} key={mask_api_key(api_key)}"
        )

        try:
            client = get_worker_client(api_key)
            return process_single_image(
                client=client,
                model=model,
                image_path=crop_path,
                logger=logger,
                rate_limiter=rate_limiter,
            )
        except Exception as exc:
            last_error = exc
            logger.log(
                f"gemini_key_failed crop={crop_path.name} "
                f"key_index={key_index + 1} key={mask_api_key(api_key)} "
                f"error={compact_error_message(exc)}"
            )
            key_manager.rotate_after_failure(
                failed_index=key_index,
                crop_name=crop_path.name,
                error=exc,
            )

    error_text = compact_error_message(last_error) if last_error is not None else "unknown error"
    raise RuntimeError(
        f"All Gemini API keys failed for crop {crop_path.name}. "
        f"attempted_keys={len(attempted_indices)} last_error={error_text}"
    ) from last_error


def crop_order_key(path: Path) -> Tuple[int, str]:
    match = re.search(r"-(\d+)$", path.stem)
    if match:
        return int(match.group(1)), path.name
    return 10**9, path.name


def extract_crop_paths(inference_result: dict) -> List[Path]:
    crop_paths: List[Path] = []
    for detection in inference_result.get("detections", []):
        crop_path = detection.get("crop_path")
        if not crop_path:
            continue
        path = Path(crop_path)
        if path.exists() and path.is_file():
            crop_paths.append(path)
    return sorted(crop_paths, key=crop_order_key)


def main() -> None:
    load_dotenv()

    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
    gemini_api_keys = parse_gemini_api_keys()
    crop_workers = parse_int_env("CROP_WORKERS", DEFAULT_CROP_WORKERS, minimum=1)
    gemini_max_concurrent = parse_int_env(
        "GEMINI_MAX_CONCURRENT_REQUESTS",
        default=crop_workers,
        minimum=1,
    )
    gemini_max_requests_per_minute = parse_int_env(
        "GEMINI_MAX_REQUESTS_PER_MINUTE",
        DEFAULT_GEMINI_MAX_REQUESTS_PER_MINUTE,
        minimum=0,
    )

    if not gemini_api_keys:
        raise RuntimeError(
            "No Gemini API keys found. Set GEMINI_API_KEY1, GEMINI_API_KEY2, ... in .env "
            "(GEMINI_API_KEY is also accepted as fallback)."
        )

    run_logger = RunLogger.create(logs_dir=LOGS_DIR, prefix="final")
    run_logger.log(f"input_images_folder={INPUT_IMAGES_FOLDER}")
    run_logger.log(f"roboflow_output_root={ROBOFLOW_OUTPUT_ROOT}")
    run_logger.log(f"output_excel_path={OUTPUT_EXCEL_PATH}")
    run_logger.log(f"model={model}")
    run_logger.log(f"gemini_key_pool_size={len(gemini_api_keys)}")
    run_logger.log(
        "gemini_keys="
        + ", ".join(
            f"{idx + 1}:{mask_api_key(key)}" for idx, key in enumerate(gemini_api_keys)
        )
    )
    run_logger.log(f"crop_workers={crop_workers}")
    run_logger.log(f"gemini_max_concurrent_requests={gemini_max_concurrent}")
    run_logger.log(
        f"gemini_max_requests_per_minute={gemini_max_requests_per_minute or 'unlimited'}"
    )
    print(f"Logs: {run_logger.file_path}")

    ensure_roboflow_dependencies()
    roboflow_settings = build_roboflow_settings()
    roboflow_model = load_roboflow_model(roboflow_settings)
    key_manager = GeminiKeyManager(api_keys=gemini_api_keys, logger=run_logger)
    rate_limiter = ApiRateLimiter(
        max_concurrent_requests=gemini_max_concurrent,
        max_requests_per_minute=gemini_max_requests_per_minute,
    )
    run_logger.log(f"roboflow_confidence={roboflow_settings.confidence}")
    run_logger.log(f"roboflow_overlap={roboflow_settings.overlap}")
    run_logger.log("roboflow_model_reuse=enabled")

    source_image_paths = get_image_paths(INPUT_IMAGES_FOLDER)
    output_dir = Path(ROBOFLOW_OUTPUT_ROOT).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    processed_source_images = 0
    success_count = 0

    try:
        with ThreadPoolExecutor(max_workers=crop_workers, thread_name_prefix="crop-worker") as pool:
            for source_image_path in source_image_paths:
                try:
                    run_logger.section(f"SOURCE IMAGE START: {source_image_path.name}")
                    inference_result = run_inference(
                        image_path=source_image_path.resolve(),
                        output_dir=output_dir,
                        settings=roboflow_settings,
                        model=roboflow_model,
                    )
                    processed_source_images += 1
                except Exception as exc:
                    run_logger.section(f"ROBOFLOW FAILED: {source_image_path.name}")
                    run_logger.log(f"error={exc}")
                    print(f"Failed Roboflow for {source_image_path.name}: {exc}")
                    continue

                result_dir = inference_result["result_dir"]
                crop_paths = extract_crop_paths(inference_result)

                run_logger.log(f"roboflow_output_dir={result_dir}")
                run_logger.log(f"roboflow_crop_count={len(crop_paths)}")
                run_logger.log_json(
                    f"ROBOFLOW DETECTIONS: {source_image_path.name}",
                    inference_result.get("detections", []),
                )

                if not crop_paths:
                    message = (
                        f"No cropped records for {source_image_path.name}. "
                        f"Output folder: {result_dir}"
                    )
                    run_logger.log(message)
                    print(message)
                    continue

                print(
                    f"Roboflow output for {source_image_path.name}: {result_dir} "
                    f"({len(crop_paths)} crop image(s))"
                )

                future_map = {
                    pool.submit(
                        process_crop_record,
                        crop_path=crop_path,
                        key_manager=key_manager,
                        model=model,
                        logger=run_logger,
                        rate_limiter=rate_limiter,
                    ): (index, crop_path)
                    for index, crop_path in enumerate(crop_paths)
                }
                pending_results: Dict[int, Tuple[bool, object]] = {}
                next_write_index = 0

                for future in as_completed(future_map):
                    crop_index, crop_path = future_map[future]
                    try:
                        record = future.result()
                        pending_results[crop_index] = (True, record)
                    except Exception as exc:
                        pending_results[crop_index] = (False, exc)

                    while next_write_index in pending_results:
                        is_success, payload = pending_results.pop(next_write_index)
                        ordered_crop_path = crop_paths[next_write_index]
                        if is_success:
                            if not isinstance(payload, dict):
                                raise RuntimeError(
                                    f"Unexpected successful payload type for {ordered_crop_path.name}: "
                                    f"{type(payload).__name__}"
                                )
                            record = payload
                            result, insurance_na = calculate_outputs(record)
                            row = build_excel_row(record, result, insurance_na)
                            append_row_to_excel(OUTPUT_EXCEL_PATH, row)
                            success_count += 1
                            run_logger.log_json(
                                f"FINAL RECORD: {ordered_crop_path.name}",
                                record,
                            )
                            print(f"Processed {ordered_crop_path.name}")
                        else:
                            exc = payload if isinstance(payload, Exception) else RuntimeError(str(payload))
                            run_logger.section(f"CROP FAILED: {ordered_crop_path.name}")
                            run_logger.log(f"error={exc}")
                            print(f"Failed {ordered_crop_path.name}: {exc}")
                        next_write_index += 1
    finally:
        run_logger.log(
            f"roboflow_processed_source_images={processed_source_images}/{len(source_image_paths)}"
        )
        run_logger.log(f"processed_cropped_record_images={success_count}")
        run_logger.log(f"output_excel={OUTPUT_EXCEL_PATH}")
        run_logger.close()

    print(
        f"Roboflow processed {processed_source_images}/{len(source_image_paths)} "
        f"source image(s)."
    )
    print(f"Processed {success_count} cropped record image(s).")
    print(f"Excel updated at: {OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    main()
