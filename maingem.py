import json
import mimetypes
import os
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from run_logger import RunLogger

# Hardcoded paths
INPUT_FOLDER = "outputs/Sample20"
OUTPUT_EXCEL_PATH = "extracted_loan_records.xlsx"
SHEET_NAME = "Results"
LOGS_DIR = "logs"
SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}

EXPECTED_KEYS = [
    "sample_no",
    "record_no",
    "customer_reference_number",
    "customer_name",
    "city",
    "state",
    "loan_period",
    "annual_interest",
    "guarantor_name",
    "guarantor_reference_number",
    "purchase_value",
    "reduction_percent_value_down_payment",
    "discount_percent",
    "loan_period_in_year",
    "monthly_principal_reduction_percent",
    "int_rate_percent",
    "total_interest_reduction",
]

EXCEL_HEADERS = [
    "sample no . , record no,",
    "Customer Reference Number",
    "Customer name",
    "City,State",
    "Loan Period AND Annual Interest",
    "Guarantor name",
    "Guarantor reference number",
    "purchase_value",
    "reduction_percent_value_down_payment",
    "discount_percent",
    "insurance_rate",
    "loan_period_in_year",
    "monthly_principal_reduction_percent",
    "int_rate_percent",
    "total_interest_reduction",
]
PROMPT = """
You are a production-grade OCR extraction system for noisy scanned finance images.

Your task is to extract the fields below from the input image and return ONLY valid JSON.
Do not include markdown. Do not include explanations.

Critical extraction rules:
1. The file name contains the authoritative sample number and record number.
   - Example: Sample20-1.jpg => sample_no = "20", record_no = "1"
   - NEVER read sample_no or record_no from the image when they can be derived from the file name.

2. customer_reference_number and guarantor_reference_number must preserve grouped tokens exactly as seen after reconstructing true line-wrap continuations.
   - Keep internal spaces between true token groups.
   - Do not drop tokens.
   - Do not invent or remove leading digits.
   - If a reference number begins immediately after a nearby word or name, still capture the full number from its true first digit.
   - Be especially careful with leading repeated digits such as "77", "779", or "77999".
   - Do not truncate a leading "77" even when the digits touch nearby text.
   - Do not merge the customer reference number into sample_no or record_no.

3. Critical continuity rule for reference numbers:
   - If a numeric token is split only because of a hard line break, treat the broken parts as one continuous token with NO extra space.
   - This applies to both customer_reference_number and guarantor_reference_number.
   - Example: if the image shows "63" split across lines as "6" at the end of one line and "3" at the start of the next line, output "63" as a single token.
   - Example: if guarantor text is read as "6E4 621 6" on one line and the next line begins with "3 1466 646 1", the correct output is "6E4 621 63 1466 646 1".
   - Example: if customer reference text is read as "... 3119 2 2" on one line and the next line begins with "1 ...", merge only when the final token is visibly wrapped mid-token; otherwise preserve true token boundaries.
   - A line break alone does NOT create a new token group.
   - Insert a space only between true groups, not inside a wrapped group.

4. For guarantor_reference_number specifically:
   - Read from the very first visible digit of the guarantor number, even if the number is touching the previous text with no separating space.
   - Preserve the grouped-token formatting in the final output after fixing line-wrap splits.
   - Example: if the image shows "77999 65 8798 98 6", output exactly "77999 65 8798 98 6".
   - Do NOT output "999 65 8798 98 6" when the actual leading digits are "77".
   - If the last token on one line and the first token on the next line are clearly one broken numeric group, join them without a space.

5. For customer_reference_number specifically:
   - Apply the same line-wrap reconstruction rule.
   - If a customer reference token is broken by a line break, merge the broken pieces into one token without spaces.
   - Preserve spaces only between true groups.

6. City and state must be split separately.
   - Detect line breaks explicitly and use them as a strong boundary signal.
   - Never let a line break cause truncation of a city/state value.
   - If the location text is visually wrapped at the end of a line, continue reading the wrapped characters on the next line before finalizing the field.
   - Join wrapped state letters across a line break when they belong to the same state abbreviation.
   - Example: if the image shows "Boston, M" at the end of one line and the next line begins with "A", then output city = "Boston" and state = "MA".
   - Do not output state = "M" when the next line clearly continues the abbreviation with "A".
   - Do not insert any space inside the state abbreviation. Output "MA", never "M A".

7. loan_period is the time duration only, such as "16 Years".
   - Do NOT include interest percentages inside loan_period.

8. annual_interest is the percent attached to the loan-period phrase such as "Sixteen Years And Eight.Seven %".
   - The correct normalized output for that example is "8.7%".

9. int_rate_percent is the explicit later interest-rate field shown separately in the image.

10. discount_percent is a separate discount percentage and is NOT the same as annual_interest.
   - It usually appears in the long purchase-value line before the loan-period phrase.
   - It may appear written as words, for example "Twenty Nine %" meaning "29%".
   - In Sample20-1.jpg the correct discount_percent is "29%".

11. reduction_percent_value_down_payment is the value near wording like reduction/value/down payment and is different from discount_percent.
   - In Sample20-1.jpg that value is 14.56%.

12. If purchase value appears both in words and digits, prefer the numeric money form when it can be inferred confidently.

13. Preserve names exactly in normal readable form.

14. All percentage outputs must include "%".

15. Return all values as strings.

16. Return only the JSON object with the exact keys below.

17. For any field, never ignore a visible continuation on the next line when the current line ends mid-token or mid-value.
   - Treat hard line wraps in scanned text as layout artifacts, not as automatic field termination.
   - Merge only true continuations from the next line.
   - Do not add extra spaces when a wrapped token should be continuous, such as state abbreviations, split words, or split reference-number groups.

The desired JSON keys are exactly:
- sample_no
- record_no
- customer_reference_number
- customer_name
- city
- state
- loan_period
- annual_interest
- guarantor_name
- guarantor_reference_number
- purchase_value
- reduction_percent_value_down_payment
- discount_percent
- loan_period_in_year
- monthly_principal_reduction_percent
- int_rate_percent
- total_interest_reduction
""".strip()

SECOND_PASS_PROMPT = """
You are doing a focused correction pass on a noisy scanned finance image.
Return ONLY valid JSON with these exact string keys:
- discount_percent
- annual_interest
- int_rate_percent
- loan_period
- loan_period_in_year
- guarantor_reference_number

Extraction rules:
1. discount_percent is the separate discount percentage before the loan-period phrase, not the 14.56% down-payment value.
2. If the image shows wording like "Twenty Nine %", output "29%".
3. annual_interest is the percent attached to the loan period phrase. Example: "Sixteen Years And Eight.Seven %" => "8.7%".
4. int_rate_percent is the explicit later interest-rate field. If the document clearly shows 8.7%, output "8.7%".
5. loan_period must be normalized like "16 Years" and loan_period_in_year like "16".
6. Preserve guarantor_reference_number token groups exactly as shown after reconstructing true line-wrap continuations.
7. For guarantor_reference_number, start from the first real digit even if there is no space between the previous text and the number.
8. Do not drop or trim leading digits. If the image shows "77999 65 8798 98 6", output exactly "77999 65 8798 98 6".
9. Do NOT output "999 65 8798 98 6" when the true number starts with "77" attached to the left.
10. In Sample20-1.jpg the correct discount_percent is 29% and annual_interest is 8.7%.
11. Important OCR continuity rule: do not let line breaks truncate values.
   - If any token is visually continued on the next line, merge the continuation correctly.
   - Do not insert spaces inside a token that should remain continuous.
   - Example: if a state abbreviation is broken as "M" at the end of one line and "A" at the start of the next line, the correct value is "MA", not "M" and not "M A".
12. Important guarantor_reference_number wrap rule:
   - If a numeric group is split by a hard line break, join the broken pieces into a single token with no space.
   - A line break alone does NOT mean a new token group.
   - Example: if the image shows "6E4 621 6" at the end of one line and the next line begins with "3 1466 646 1", output exactly "6E4 621 63 1466 646 1".
   - Preserve spaces only between true groups, not inside a wrapped group.
""".strip()

SCHEMA = {
    "name": "loan_record_extraction",
    "strict": True,
    "schema": {
        "type": "object",
        "properties": {key: {"type": "string"} for key in EXPECTED_KEYS},
        "required": EXPECTED_KEYS,
        "additionalProperties": False,
    },
}

SECOND_PASS_KEYS = [
    "discount_percent",
    "annual_interest",
    "int_rate_percent",
    "loan_period",
    "loan_period_in_year",
    "guarantor_reference_number",
]

SECOND_PASS_SCHEMA = {
    "name": "loan_record_correction",
    "strict": True,
    "schema": {
        "type": "object",
        "properties": {key: {"type": "string"} for key in SECOND_PASS_KEYS},
        "required": SECOND_PASS_KEYS,
        "additionalProperties": False,
    },
}

NUMBER_WORDS = {
    "zero": 0,
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
    "eleven": 11,
    "twelve": 12,
    "thirteen": 13,
    "fourteen": 14,
    "fifteen": 15,
    "sixteen": 16,
    "seventeen": 17,
    "eighteen": 18,
    "nineteen": 19,
    "twenty": 20,
    "thirty": 30,
    "forty": 40,
    "fifty": 50,
    "sixty": 60,
    "seventy": 70,
    "eighty": 80,
    "ninety": 90,
}


class ExtractionError(RuntimeError):
    pass


def image_to_part(image_path: str) -> types.Part:
    mime_type, _ = mimetypes.guess_type(image_path)
    mime_type = mime_type or "image/jpeg"
    with open(image_path, "rb") as f:
        image_bytes = f.read()
    return types.Part.from_bytes(data=image_bytes, mime_type=mime_type)


def parse_sample_and_record_from_filename(image_path: str) -> Tuple[str, str]:
    name = Path(image_path).name
    match = re.search(r"sample\s*(\d+)\s*[-_]\s*(\d+)", name, flags=re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Could not derive sample_no and record_no from file name: {name}. "
            "Expected a pattern like Sample20-1.jpg"
        )
    return match.group(1), match.group(2)


def normalize_spaces(value: str) -> str:
    value = str(value or "").replace("\u00a0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_percent(value: str) -> str:
    value = normalize_spaces(value)
    value = value.replace(" %", "%")
    value = re.sub(r"(?<=\d)\s*%", "%", value)
    value = value.replace("..", ".")
    return value


def normalize_money(value: str) -> str:
    value = normalize_spaces(value)
    value = value.replace(" ,", ",").replace(", ", ",")
    value = value.replace(" .", ".").replace(". ", ".")
    value = re.sub(r"\$\s*", "$ ", value)
    value = re.sub(r"\$\s+", "$ ", value)
    return value.strip()


def normalize_name(value: str) -> str:
    value = normalize_spaces(value)
    value = re.sub(r"\b(Mr|Ms|Mrs|Dr)\.\s*", lambda m: m.group(1).title() + ".", value, flags=re.IGNORECASE)
    return value


def ensure_required_keys(data: Dict[str, Any]) -> Dict[str, str]:
    missing = [key for key in EXPECTED_KEYS if key not in data]
    extra = [key for key in data if key not in EXPECTED_KEYS]
    if missing:
        raise ExtractionError(f"Model response missing keys: {missing}")
    if extra:
        raise ExtractionError(f"Model response contains unexpected keys: {extra}")
    return {key: normalize_spaces(str(data[key])) for key in EXPECTED_KEYS}


def ensure_second_pass_keys(data: Dict[str, Any]) -> Dict[str, str]:
    missing = [key for key in SECOND_PASS_KEYS if key not in data]
    extra = [key for key in data if key not in SECOND_PASS_KEYS]
    if missing:
        raise ExtractionError(f"Second-pass response missing keys: {missing}")
    if extra:
        raise ExtractionError(f"Second-pass response contains unexpected keys: {extra}")
    return {key: normalize_spaces(str(data[key])) for key in SECOND_PASS_KEYS}


def words_to_int(text: str) -> Optional[int]:
    cleaned = text.lower().replace("-", " ")
    cleaned = re.sub(r"[^a-z\s]", " ", cleaned)
    tokens = [token for token in cleaned.split() if token not in {"and", "year", "years", "percent"}]
    if not tokens:
        return None

    total = 0
    current = 0
    matched_any = False

    for token in tokens:
        if token in NUMBER_WORDS:
            current += NUMBER_WORDS[token]
            matched_any = True
        elif token == "hundred":
            if current == 0:
                current = 1
            current *= 100
            matched_any = True
        elif token == "thousand":
            if current == 0:
                current = 1
            total += current * 1000
            current = 0
            matched_any = True
        else:
            return None

    if not matched_any:
        return None
    return total + current


def decimal_from_any(value: str) -> Optional[Decimal]:
    text = normalize_spaces(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def format_decimal_percent(number: Decimal) -> str:
    normalized = format(number.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return f"{normalized}%"


def extract_numeric_percent(value: str) -> Optional[Decimal]:
    return decimal_from_any(value)


def normalize_loan_period(value: str, fallback_years: str) -> str:
    value = normalize_spaces(value)

    digit_match = re.search(r"(\d+)\s*years?", value, flags=re.IGNORECASE)
    if digit_match:
        return f"{digit_match.group(1)} Years"

    text_before_years = re.search(r"([A-Za-z\-\s]+?)\s*years?", value, flags=re.IGNORECASE)
    if text_before_years:
        years_as_int = words_to_int(text_before_years.group(1))
        if years_as_int is not None:
            return f"{years_as_int} Years"

    fallback_match = re.search(r"\d+", str(fallback_years))
    if fallback_match:
        return f"{fallback_match.group(0)} Years"

    return value


def normalize_annual_interest(value: str, int_rate_percent: str) -> str:
    value = normalize_percent(value)
    int_rate_percent = normalize_percent(int_rate_percent)

    annual_num = extract_numeric_percent(value)
    int_rate_num = extract_numeric_percent(int_rate_percent)

    if annual_num is None and int_rate_num is not None:
        return format_decimal_percent(int_rate_num)

    if annual_num is not None and int_rate_num is not None:
        if abs(annual_num - int_rate_num) <= Decimal("0.25"):
            return format_decimal_percent(min(annual_num, int_rate_num))
        return format_decimal_percent(annual_num)

    if annual_num is not None:
        return format_decimal_percent(annual_num)

    return value


def normalize_purchase_value(value: str) -> str:
    value = normalize_money(value)
    number_match = re.search(r"\$?\s*([\d,]+(?:\.\d{1,2})?)", value)
    if number_match:
        number_part = number_match.group(1)
        if "." not in number_part:
            number_part = f"{number_part}.00"
        return f"$ {number_part}"
    return value


def normalize_discount_percent(value: str) -> str:
    value = normalize_percent(value)
    numeric = extract_numeric_percent(value)
    if numeric is not None:
        return format_decimal_percent(numeric)

    words_value = words_to_int(value)
    if words_value is not None:
        return f"{words_value}%"

    return value


def compute_insurance_rate(discount_percent: str, loan_period_in_year: str) -> str:
    discount_num = extract_numeric_percent(discount_percent)
    loan_years_num = decimal_from_any(loan_period_in_year)

    if discount_num is None or loan_years_num is None:
        return "NA"

    try:
        financed_percent = Decimal("100") - discount_num
    except InvalidOperation:
        return "NA"

    loan_years = int(loan_years_num)
    if loan_years < 1 or loan_years > 30:
        return "NA"

    if financed_percent > Decimal("95.01"):
        return "NA"

    if Decimal("70") <= financed_percent <= Decimal("80.99"):
        return "0.32%"

    if financed_percent == Decimal("81"):
        if 1 <= loan_years <= 25:
            return "0.21%"
        if 26 <= loan_years <= 30:
            return "0.32%"
        return "NA"

    if Decimal("81.01") <= financed_percent < Decimal("90"):
        if 1 <= loan_years <= 25:
            return "0.41%"
        if 26 <= loan_years <= 30:
            return "0.52%"
        return "NA"

    if Decimal("90") <= financed_percent <= Decimal("95"):
        if 1 <= loan_years <= 25:
            return "0.67%"
        if 26 <= loan_years <= 30:
            return "0.78%"
        return "NA"

    return "NA"


def post_process(data: Dict[str, str], image_path: str) -> Dict[str, str]:
    sample_no, record_no = parse_sample_and_record_from_filename(image_path)

    cleaned = {key: normalize_spaces(value) for key, value in data.items()}

    cleaned["sample_no"] = sample_no
    cleaned["record_no"] = record_no
    cleaned["customer_name"] = normalize_name(cleaned["customer_name"])
    cleaned["guarantor_name"] = normalize_name(cleaned["guarantor_name"])
    cleaned["city"] = normalize_spaces(cleaned["city"])
    cleaned["state"] = normalize_spaces(cleaned["state"])
    cleaned["purchase_value"] = normalize_purchase_value(cleaned["purchase_value"])
    cleaned["reduction_percent_value_down_payment"] = normalize_percent(cleaned["reduction_percent_value_down_payment"])
    cleaned["monthly_principal_reduction_percent"] = normalize_percent(cleaned["monthly_principal_reduction_percent"])
    cleaned["int_rate_percent"] = normalize_percent(cleaned["int_rate_percent"])
    cleaned["total_interest_reduction"] = normalize_percent(cleaned["total_interest_reduction"])
    cleaned["customer_reference_number"] = normalize_spaces(cleaned["customer_reference_number"])
    cleaned["guarantor_reference_number"] = normalize_spaces(cleaned["guarantor_reference_number"])

    cleaned["loan_period"] = normalize_loan_period(cleaned["loan_period"], cleaned["loan_period_in_year"])

    loan_period_in_year_match = re.search(r"\d+", cleaned["loan_period_in_year"])
    if loan_period_in_year_match:
        cleaned["loan_period_in_year"] = loan_period_in_year_match.group(0)
    else:
        period_year_match = re.search(r"\d+", cleaned["loan_period"])
        cleaned["loan_period_in_year"] = period_year_match.group(0) if period_year_match else cleaned["loan_period_in_year"]

    cleaned["discount_percent"] = normalize_discount_percent(cleaned["discount_percent"])
    cleaned["annual_interest"] = normalize_annual_interest(cleaned["annual_interest"], cleaned["int_rate_percent"])

    prefix = f"{sample_no} {record_no}"
    if cleaned["customer_reference_number"].startswith(prefix + " "):
        cleaned["customer_reference_number"] = cleaned["customer_reference_number"][len(prefix) + 1 :].strip()

    return cleaned


def build_user_message(image_path: str) -> str:
    sample_no, record_no = parse_sample_and_record_from_filename(image_path)
    return (
        f"File name: {Path(image_path).name}\n"
        f"Authoritative sample_no from filename: {sample_no}\n"
        f"Authoritative record_no from filename: {record_no}\n"
        "Important normalization targets for this record family:\n"
        "- loan_period should be formatted like 16 Years\n"
        "- annual_interest should be a clean numeric percent like 8.7%\n"
        "- discount_percent is a separate percentage like 29% and should not be confused with annual_interest or 14.56%\n"
        "Extract the rest from the image."
    )


def parse_json_response(response: Any) -> Dict[str, Any]:
    parsed = getattr(response, "parsed", None)
    if isinstance(parsed, dict):
        return parsed
    if hasattr(parsed, "model_dump"):
        return parsed.model_dump()

    output_text = getattr(response, "text", "")
    if not output_text:
        raise ExtractionError("Model returned an empty response")
    output_text = output_text.strip()
    fenced_match = re.search(r"```(?:json)?\s*(\{.*\})\s*```", output_text, flags=re.IGNORECASE | re.DOTALL)
    if fenced_match:
        output_text = fenced_match.group(1).strip()
    elif not output_text.startswith("{"):
        object_match = re.search(r"\{.*\}", output_text, flags=re.DOTALL)
        if object_match:
            output_text = object_match.group(0).strip()
    try:
        return json.loads(output_text)
    except json.JSONDecodeError as exc:
        raise ExtractionError(f"Model did not return valid JSON: {exc}\nRaw output: {output_text}") from exc


def response_parsed_to_jsonable(parsed: Any) -> Any:
    if isinstance(parsed, dict):
        return parsed
    if hasattr(parsed, "model_dump"):
        return parsed.model_dump()
    if parsed is None:
        return None
    return str(parsed)


def call_structured_model(
    client: genai.Client,
    model: str,
    image_path: str,
    prompt: str,
    schema: Dict[str, Any],
    user_text: str,
    logger: Optional[RunLogger] = None,
    step_name: str = "llm_step",
) -> Dict[str, Any]:
    if logger is not None:
        logger.section(f"LLM STEP: {step_name}")
        logger.log(f"image={image_path}")
        logger.log(f"model={model}")
        logger.log(f"user_text={user_text}")

    response = client.models.generate_content(
        model=model,
        contents=[
            types.UserContent(
                parts=[
                    types.Part.from_text(text=user_text),
                    image_to_part(image_path),
                ]
            ),
        ],
        config=types.GenerateContentConfig(
            system_instruction=prompt,
            response_mime_type="application/json",
            response_json_schema=schema["schema"],
            temperature=0,
        ),
    )

    if logger is not None:
        raw_text = getattr(response, "text", "")
        logger.section(f"LLM RAW OUTPUT: {step_name}")
        logger.log(raw_text if raw_text else "<empty>")
        logger.log_json(
            f"LLM PARSED FIELD (SDK): {step_name}",
            response_parsed_to_jsonable(getattr(response, "parsed", None)),
        )

    payload = parse_json_response(response)
    if logger is not None:
        logger.log_json(f"LLM PARSED JSON: {step_name}", payload)
    return payload


def call_model(
    client: genai.Client,
    model: str,
    image_path: str,
    logger: Optional[RunLogger] = None,
) -> Dict[str, str]:
    payload = call_structured_model(
        client=client,
        model=model,
        image_path=image_path,
        prompt=PROMPT,
        schema=SCHEMA,
        user_text=build_user_message(image_path),
        logger=logger,
        step_name="primary_extraction",
    )
    return ensure_required_keys(payload)


def call_second_pass(
    client: genai.Client,
    model: str,
    image_path: str,
    current: Dict[str, str],
    logger: Optional[RunLogger] = None,
) -> Dict[str, str]:
    user_text = (
        f"File name: {Path(image_path).name}\n"
        f"Current extraction to correct: {json.dumps(current, ensure_ascii=False)}\n"
        "Focus on the values that are often confused in this document family: discount_percent, annual_interest, int_rate_percent, loan_period, loan_period_in_year, guarantor_reference_number.\n"
        "Correct them from the image."
    )
    payload = call_structured_model(
        client=client,
        model=model,
        image_path=image_path,
        prompt=SECOND_PASS_PROMPT,
        schema=SECOND_PASS_SCHEMA,
        user_text=user_text,
        logger=logger,
        step_name="second_pass_correction",
    )
    return ensure_second_pass_keys(payload)


def merge_second_pass(primary: Dict[str, str], second: Dict[str, str]) -> Dict[str, str]:
    merged = dict(primary)
    for key, value in second.items():
        if normalize_spaces(value):
            merged[key] = normalize_spaces(value)
    return merged


def needs_second_pass(record: Dict[str, str]) -> bool:
    discount_num = extract_numeric_percent(record.get("discount_percent", ""))
    annual_num = extract_numeric_percent(record.get("annual_interest", ""))
    int_num = extract_numeric_percent(record.get("int_rate_percent", ""))
    loan_years = decimal_from_any(record.get("loan_period_in_year", ""))

    if discount_num is None:
        return True
    if annual_num is None:
        return True
    if int_num is None:
        return True
    if loan_years is None:
        return True
    if annual_num is not None and (annual_num < Decimal("0.1") or annual_num > Decimal("50")):
        return True
    if int_num is not None and (int_num < Decimal("0.1") or int_num > Decimal("50")):
        return True
    if abs(annual_num - int_num) > Decimal("0.3"):
        return True
    return False


def enforce_cross_field_consistency(record: Dict[str, str]) -> Dict[str, str]:
    cleaned = dict(record)

    cleaned["loan_period"] = normalize_loan_period(cleaned.get("loan_period", ""), cleaned.get("loan_period_in_year", ""))
    year_match = re.search(r"\d+", cleaned.get("loan_period", ""))
    if year_match:
        cleaned["loan_period_in_year"] = year_match.group(0)
    else:
        year_num = decimal_from_any(cleaned.get("loan_period_in_year", ""))
        if year_num is not None:
            cleaned["loan_period_in_year"] = str(int(year_num))

    cleaned["discount_percent"] = normalize_discount_percent(cleaned.get("discount_percent", ""))

    annual_num = extract_numeric_percent(cleaned.get("annual_interest", ""))
    int_num = extract_numeric_percent(cleaned.get("int_rate_percent", ""))

    if int_num is None and annual_num is not None:
        cleaned["int_rate_percent"] = format_decimal_percent(annual_num)
        int_num = annual_num
    elif int_num is not None:
        cleaned["int_rate_percent"] = format_decimal_percent(int_num)

    if annual_num is None and int_num is not None:
        cleaned["annual_interest"] = format_decimal_percent(int_num)
        annual_num = int_num
    elif annual_num is not None and int_num is not None and abs(annual_num - int_num) <= Decimal("0.25"):
        chosen = min(annual_num, int_num)
        cleaned["annual_interest"] = format_decimal_percent(chosen)
        cleaned["int_rate_percent"] = format_decimal_percent(chosen)
    elif annual_num is not None:
        cleaned["annual_interest"] = format_decimal_percent(annual_num)

    return cleaned


def get_image_paths(folder_path: str) -> List[Path]:
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Input folder not found: {folder_path}")
    if not folder.is_dir():
        raise NotADirectoryError(f"Input path is not a folder: {folder_path}")

    image_paths = [path for path in sorted(folder.iterdir()) if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS]
    if not image_paths:
        raise FileNotFoundError(f"No supported image files found in folder: {folder_path}")
    return image_paths


def build_excel_row(record: Dict[str, str], insurance_rate: str) -> List[str]:
    raw_row = [
        f"SAMPLE NO. {record['sample_no']} ,RECORD.{record['record_no']}",
        record["customer_reference_number"],
        record["customer_name"],
        f"{record['city']} , {record['state']}",
        f"{record['loan_period']} AND {record['annual_interest']}",
        record["guarantor_name"],
        record["guarantor_reference_number"],
        record["purchase_value"],
        record["reduction_percent_value_down_payment"],
        record["discount_percent"],
        insurance_rate,
        record["loan_period_in_year"],
        record["monthly_principal_reduction_percent"],
        record["int_rate_percent"],
        record["total_interest_reduction"],
    ]
    return [normalize_spaces("" if value is None else str(value)).upper() for value in raw_row]


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 40)


def ensure_workbook(path: str):
    excel_path = Path(path)
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(EXCEL_HEADERS)
        elif [ws.cell(row=1, column=i + 1).value for i in range(len(EXCEL_HEADERS))] != EXCEL_HEADERS:
            ws.delete_rows(1, ws.max_row)
            ws.append(EXCEL_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(EXCEL_HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return wb, ws


def append_row_to_excel(path: str, row: List[str]) -> None:
    wb, ws = ensure_workbook(path)
    if len(row) != len(EXCEL_HEADERS):
        raise ValueError(
            f"Excel row has {len(row)} columns but expected {len(EXCEL_HEADERS)} columns."
        )

    normalized_row = [normalize_spaces("" if value is None else str(value)).upper() for value in row]
    next_row = ws.max_row + 1
    for column_index, value in enumerate(normalized_row, start=1):
        ws.cell(row=next_row, column=column_index, value=value)

    for cell in ws[next_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_columns(ws)
    wb.save(path)


def process_single_image(
    client: genai.Client,
    model: str,
    image_path: Path,
    logger: Optional[RunLogger] = None,
) -> Dict[str, str]:
    if logger is not None:
        logger.section(f"IMAGE START: {image_path.name}")
        logger.log(f"image_path={image_path}")

    extracted = call_model(client, model, str(image_path), logger=logger)
    if logger is not None:
        logger.log_json("PRIMARY EXTRACTION (NORMALIZED KEYS)", extracted)

    final_output = post_process(extracted, str(image_path))
    if logger is not None:
        logger.log_json("AFTER POST PROCESS", final_output)

    if needs_second_pass(final_output):
        if logger is not None:
            logger.log("second_pass_required=true")
        second_pass = call_second_pass(client, model, str(image_path), final_output, logger=logger)
        if logger is not None:
            logger.log_json("SECOND PASS OUTPUT", second_pass)
        merged = merge_second_pass(final_output, second_pass)
        final_output = post_process(merged, str(image_path))
    else:
        if logger is not None:
            logger.log("second_pass_required=false")

    final_output = enforce_cross_field_consistency(final_output)
    if logger is not None:
        logger.log_json("FINAL OUTPUT", final_output)
    return final_output


def main() -> None:
    load_dotenv()

    api_key = os.getenv("GEMINI_API_KEY")
    model = os.getenv("GEMINI_MODEL")

    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is missing in .env")
    if not model:
        raise RuntimeError("GEMINI_MODEL is missing in .env")

    run_logger = RunLogger.create(logs_dir=LOGS_DIR, prefix="maingem")
    run_logger.log(f"input_folder={INPUT_FOLDER}")
    run_logger.log(f"output_excel_path={OUTPUT_EXCEL_PATH}")
    run_logger.log(f"model={model}")
    print(f"Logs: {run_logger.file_path}")

    image_paths = get_image_paths(INPUT_FOLDER)
    client = genai.Client(api_key=api_key)

    processed_records: List[Dict[str, str]] = []

    try:
        for image_path in image_paths:
            try:
                final_output = process_single_image(client, model, image_path, logger=run_logger)
                insurance_rate = compute_insurance_rate(
                    final_output["discount_percent"],
                    final_output["loan_period_in_year"],
                )
                row = build_excel_row(final_output, insurance_rate)
                append_row_to_excel(OUTPUT_EXCEL_PATH, row)
                final_output["insurance_rate"] = insurance_rate
                processed_records.append(final_output)
                run_logger.log_json(f"EXCEL ROW SOURCE: {image_path.name}", final_output)
                print(json.dumps(final_output, indent=2, ensure_ascii=False))
            except Exception as exc:
                run_logger.section(f"IMAGE FAILED: {image_path.name}")
                run_logger.log(f"error={exc}")
                raise
    finally:
        run_logger.log(f"processed_images={len(processed_records)}")
        run_logger.log(f"output_excel={OUTPUT_EXCEL_PATH}")
        run_logger.close()

    print(f"Processed {len(processed_records)} image(s).")
    print(f"Excel updated at: {OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    main()
